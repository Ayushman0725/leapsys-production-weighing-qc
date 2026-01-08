import os, time, serial
import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
from datetime import datetime
from PIL import Image, ImageTk, ImageEnhance, ImageFilter

# ================= CONFIG =================
PORT = "COM3"
BAUD = 9600
TOLERANCE = 5  # ± grams allowed

SAVE_FOLDER = r"C:\weight data"
os.makedirs(SAVE_FOLDER, exist_ok=True)
FILE_PATH = os.path.join(SAVE_FOLDER, "leapsys_weighpro.xlsx")

BG_IMAGE = r"C:\Users\aarya\Downloads\antique-weighing-scale_92242-183.jpg"

# ================= SERIAL =================
try:
    ser = serial.Serial(PORT, BAUD, timeout=2)
except:
    ser = None

# ================= FIXED RECIPES =================
RECIPE_DATA = {
    "MAGGI": {"Noodles": 420, "Water": 840, "Masala": 35, "Oil": 20},
    "PASTA": {"Pasta": 500, "Water": 1000, "Sauce": 120, "Oil": 30},
    "SANDWICH": {"Bread": 120, "Butter": 30, "Filling": 150, "Sauce": 40},
    "PIZZA": {"Base": 150, "Sauce": 80, "Cheese": 120, "Toppings": 100}
}

SAMPLES = {
    "Sample Weight 1": ["MAGGI"],
    "Sample Weight 2": ["PASTA"],
    "Sample Weight 3": ["SANDWICH"],
    "Sample Weight 4": ["PIZZA"]
}

# ================= ROOT =================
root = tk.Tk()
root.title("Leapsys WeighPro")
root.geometry("1200x720")
root.resizable(False, False)

# ================= BACKGROUND =================
def set_background():
    img = Image.open(BG_IMAGE).resize((1200, 720))
    img = img.filter(ImageFilter.GaussianBlur(7))
    img = ImageEnhance.Brightness(img).enhance(1.2)
    img = ImageEnhance.Contrast(img).enhance(0.9)
    bg = ImageTk.PhotoImage(img)
    lbl = tk.Label(root, image=bg)
    lbl.image = bg
    lbl.place(x=0, y=0, relwidth=1, relheight=1)

set_background()

# ================= UTILS =================
def clear():
    for w in root.winfo_children()[1:]:
        w.destroy()

def read_weight(label):
    if ser is None:
        messagebox.showerror("Scale Error", "CMO-3 not connected")
        return

    ser.reset_input_buffer()
    time.sleep(0.2)

    for _ in range(60):
        line = ser.readline().decode(errors="ignore")
        cleaned = ''.join(c if c in "0123456789.+-" else " " for c in line)

        for part in cleaned.split():
            try:
                val = float(part)
                if val < 100:
                    val *= 1000
                label.config(text=f"{val:.2f}")
                return
            except:
                pass

    messagebox.showerror("Scale Error", "No weight detected")

def save_data(sample, recipe, ing, target, actual):
    if actual["text"] == "--":
        return

    actual_val = float(actual["text"])
    status = "PASS" if abs(actual_val - target) <= TOLERANCE else "FAIL"

    if not os.path.exists(FILE_PATH):
        wb = Workbook()
        ws = wb.active
        ws.append([
            "DateTime", "Sample", "Recipe",
            "Ingredient", "Target(g)", "Actual(g)", "Status"
        ])
    else:
        wb = load_workbook(FILE_PATH)
        ws = wb.active

    ws.append([
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        sample, recipe, ing, target, actual_val, status
    ])
    wb.save(FILE_PATH)

    messagebox.showinfo("Saved", f"{ing} saved ({status})")

# ================= PRODUCTION SUMMARY =================
def production_summary():
    clear()

    tk.Label(root, text="PRODUCTION SUMMARY",
             font=("Segoe UI", 32, "bold"),
             fg="#facc15").pack(pady=20)

    if not os.path.exists(FILE_PATH):
        messagebox.showinfo("Info", "No production data found")
        home()
        return

    wb = load_workbook(FILE_PATH)
    ws = wb.active
    today = datetime.now().strftime("%Y-%m-%d")

    summary = {}

    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r[0].startswith(today):
            continue

        recipe = r[2]
        actual = r[5]
        status = r[6]

        summary.setdefault(recipe, {"total": 0, "pass": 0, "fail": 0, "weights": []})
        summary[recipe]["total"] += 1
        summary[recipe]["weights"].append(actual)

        if status == "PASS":
            summary[recipe]["pass"] += 1
        else:
            summary[recipe]["fail"] += 1

    for recipe, d in summary.items():
        avg = round(sum(d["weights"]) / len(d["weights"]), 2)

        card = tk.Frame(root, bg="#020617")
        card.pack(fill="x", padx=180, pady=10)

        tk.Label(card, text=recipe,
                 font=("Segoe UI", 20, "bold"),
                 bg="#020617", fg="#38bdf8").pack(anchor="w", padx=20)

        tk.Label(card,
                 text=f"Total: {d['total']}   PASS: {d['pass']}   FAIL: {d['fail']}   Avg: {avg} g",
                 font=("Segoe UI", 16),
                 bg="#020617", fg="white").pack(anchor="w", padx=20, pady=8)

    tk.Button(root, text="⬅ Back",
              command=home,
              bg="#334155", fg="white",
              font=("Segoe UI", 14),
              relief="flat").pack(pady=25)

# ================= UI =================
def home():
    clear()

    tk.Label(root, text="WEIGHPRO",
             font=("Segoe UI", 40, "bold"),
             fg="#facc15").pack(pady=(30, 5))

    tk.Label(root, text="Powered by LEAPSYS",
             font=("Segoe UI", 14, "bold"),
             fg="#94a3b8").pack(pady=(0, 35))

    for s in SAMPLES:
        tk.Button(root, text=s,
                  width=25, height=2,
                  font=("Segoe UI", 18, "bold"),
                  bg="#0f766e", fg="white",
                  activebackground="#115e59",
                  relief="flat",
                  command=lambda x=s: sample_page(x)).pack(pady=12)

    tk.Button(root, text="📊 Production Summary",
              command=production_summary,
              bg="#1e40af", fg="white",
              font=("Segoe UI", 16, "bold"),
              relief="flat").pack(pady=20)

def sample_page(sample):
    clear()
    recipe = SAMPLES[sample][0]

    tk.Label(root, text=recipe,
             font=("Segoe UI", 30, "bold"),
             fg="#22c55e").pack(pady=20)

    for ing, target in RECIPE_DATA[recipe].items():
        row = tk.Frame(root, bg="#020617")
        row.pack(fill="x", padx=180, pady=6)

        tk.Label(row, text=ing, width=18,
                 font=("Segoe UI", 16),
                 bg="#020617", fg="white").pack(side="left")

        tk.Label(row, text=f"{target} g", width=10,
                 font=("Segoe UI", 16, "bold"),
                 bg="#020617", fg="#22c55e").pack(side="left")

        actual = tk.Label(row, text="--", width=10,
                          font=("Segoe UI", 16, "bold"),
                          bg="#020617", fg="#38bdf8")
        actual.pack(side="left", padx=5)

        tk.Button(row, text="READ",
                  bg="#2563eb", fg="white",
                  font=("Segoe UI", 14, "bold"),
                  relief="flat",
                  command=lambda a=actual: read_weight(a)).pack(side="left", padx=8)

        tk.Button(row, text="SAVE",
                  bg="#16a34a", fg="white",
                  font=("Segoe UI", 14, "bold"),
                  relief="flat",
                  command=lambda i=ing, t=target, a=actual:
                  save_data(sample, recipe, i, t, a)).pack(side="left", padx=8)

    tk.Button(root, text="⬅ Back",
              command=home,
              bg="#334155", fg="white",
              font=("Segoe UI", 14),
              relief="flat").pack(pady=25)

# ================= START =================
home()
root.mainloop()
